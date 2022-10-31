""" Autoupdater for Engineering Timing Release Plan and eBOM Export

Gathers data from Upchain eBOM exports, google sheets,
custom Upchain email reports and consolidates them into
a single excel file, maintaining the formula that already exist
in the master file. Requires BOM Exports from Upchain on a daily
basis run through the BOM Analyser VBA document and the custom
Upchain reporting that is recieved by email daily.

"""

from datetime import date, timedelta
import datetime
import os
import glob
import logging
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import xlwings as xl
import gspread

BASE_PATH = r"S:\PDM Files\P1 - Mustang\\"
TARGET_PATH = fr"{BASE_PATH}\ETRS\ETRS Master\ETRS v7 Master.xlsx"

weekday = {
    "Monday" : 0,
    "Tuesday" : 1,
    "Wednesday" : 2,
    "Thursday" : 3,
    "Friday" : 4,
    "Saturday" : 5,
    "Sunday" : 6,
}

now = datetime.datetime.today()
bom_export_path = r"\BOM\BOM Exports\BOM Export "
today = date.today()
this_monday = now + timedelta(days = (weekday["Friday"]- now.weekday()))
last_friday = now + timedelta(days = (weekday["Friday"]- now.weekday()), weeks = -1)
monday_bom_format = str(this_monday.strftime('%Y%m%d'))
today_bom_format = str(today.strftime('%Y%m%d'))
today_fin_format = str(today.strftime('%Y-%m-%d'))
yesterday_bom_format = str((today - timedelta(days=1)).strftime('%Y%m%d'))
friday_bom_format = str(last_friday.strftime('%Y%m%d'))

# Trailing space is important workflow_path!
custom_path = r"\BOM\Upchain Custom Reports"
workflow_path = fr"{custom_path}\EBOM Reports\eBOM Workflow Report "

today_bom_source = fr"{BASE_PATH}{bom_export_path}{today_bom_format}.xlsx"
yesterday_bom_source = fr"{BASE_PATH}{bom_export_path}{yesterday_bom_format}.xlsx"
monday_bom_source = fr"{BASE_PATH}{bom_export_path}{monday_bom_format}.xlsx"
friday_bom_source = fr"{BASE_PATH}{bom_export_path}{friday_bom_format}.xlsx"
workflow_source = fr"{BASE_PATH}{workflow_path}{today_bom_format}.xlsx"
complete_source = fr"{BASE_PATH}{custom_path}\Complete Report\Complete Report {today_bom_format}.csv"
purchasing_source = fr"{BASE_PATH}ETRS\DataFiles\Finance {today}.csv"
timing_source = fr"{BASE_PATH}ETRS\New Parts\Timing Sheet.xlsx"

today_bom_df = pd.read_excel(today_bom_source, sheet_name="Formatted BOM", index_col=None, na_values=["NA"], usecols="D:F")
yesterday_bom_df = pd.read_excel(yesterday_bom_source, sheet_name="Formatted BOM", index_col=None, na_values=["NA"], usecols="D:F")
monday_bom_df = pd.read_excel(monday_bom_source, sheet_name="Formatted BOM", index_col=None, na_values=["NA"], usecols="D:F")

logging.basicConfig(format='%(asctime)s:%(levelname)s:%(message)s - ',
                    encoding='utf-8',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    level=logging.DEBUG,
                    handlers=[
                        logging.FileHandler("etrs_updater.log"),
                        logging.StreamHandler()
                        ]
                    )

def connect_to_google_sheet(sheet_id):
    """ Connects to Google Sheets and returns values as a list of lists """
    logging.info("Connecting to Google Sheets...")
    google_connect = gspread.service_account(filename="export-finance-formatted-bom-6c652f359e2f.json")
    sheet = google_connect.open_by_key(sheet_id).sheet1
    logging.info("Connected")
    return sheet


def write_to_finance_update_csv(sheet_key, filename):
    """ Retrieves values from Google Sheet specified in the sheet_key

    sheet_key is a string that can be found in the URL of the target sheet you
    are connecting to. This function retrieves values from the Google Sheet
    and saves them in a CSV labelled with today's date.
    """

    logging.info("Retrieving values from Google Sheets")
    google_sheet = connect_to_google_sheet(sheet_key)
    data_frame = pd.DataFrame(google_sheet.get_all_values())

    logging.info("Writing values to CSV at %s", filename)
    data_frame.to_csv(filename, index=False, header=False)


def load_data_file(source_path):
    """ Loads data from CSV or XLSX into a dataframe

    source_path should be the file path of either a CSV or an Excel. If the
    file is a Formatted BOM, the sheet name will be correctly labelled in the
    ETRS, otherwise it will follow default naming conventions
    """

    if source_path[-4:] == "csv ":
        logging.info("CSV found, reading csv")
        data = pd.read_csv(source_path)
    elif source_path[-24:-14] == "BOM Export":
        logging.info("BOM Export found, reading excel")
        data = pd.read_excel(source_path, sheet_name="Formatted BOM")
    elif source_path[-4:] == "xlsx":
        logging.info("Excel doc found, reading excel")
        data = pd.read_excel(source_path)
    else:
        logging.critical("Document not found")

    return data


def excel_archiver():
    """ Moves any old excel documents to the Archive folder from the ETRS folder
    """
    existing_file_list = glob.glob(fr"{BASE_PATH}\ETRS\*.xlsx")
    datetime_timestamp = datetime.datetime.now()
    string_timestamp = datetime_timestamp.strftime("%H-%M-%S")

    for i in existing_file_list:
        path, file = os.path.split(i)
        existing_filename, extension = os.path.splitext(file)
        archive_path = fr"{BASE_PATH}ETRS\Archive\{existing_filename} -- {string_timestamp}{extension}"
        os.rename(i,archive_path)


def write_to_etrs():
    """ Collects the various data sources and writes them to an XLSX file

    If you have the accompanying ETRS Master file at targetPath, this export
    will conform to the requirements of that sheet to automate the creation of
    a new ETRS file and archive the one. Ideally this process happens
    on a daily basis.

    """

    logging.info("Loading ETRS Workbook %s", TARGET_PATH)
    logging.info("Note, this might take a while...")
    book = openpyxl.load_workbook(TARGET_PATH)

    with pd.ExcelWriter(TARGET_PATH, engine='openpyxl', mode='a', # pylint: disable=abstract-class-instantiated
                        if_sheet_exists="replace") as writer:
        
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        all_data_exports = {
            "BOM Export": ["Formatted BOM", today_bom_source],
            "Raw BOM" : ["Raw BOM", today_bom_source],
            "Purchasing Lead Times" : [f"Finance {today_fin_format}", purchasing_source],
            "Monday's BOM Export" : ["Formatted BOM", monday_bom_source],
            "Friday's BOM Export" : ["Formatted BOM", friday_bom_source],
            "Workflow" : ["Report Data", workflow_source],
            "Yesterday BOM Export" : ["Formatted BOM", yesterday_bom_source],
            "Complete Export" : [f"Complete Report {today_bom_format}",complete_source],
            "Timing" : ["Timing", timing_source ]
        }

        items = all_data_exports.items()

        for item in items:
            logging.info("Loading %s from source" , item[0])
            print(item[1][1])
            if os.path.exists(item[1][1]) and item[1][1][-4:] == "xlsx":
                loaded_data = pd.read_excel(item[1][1], sheet_name=item[1][0])
                loaded_data.to_excel(writer, sheet_name=item[0])
                logging.info("Writing %s to ETRS" , item[0])

            elif os.path.exists(item[1][1]) and item[1][1][-4:] == ".csv":
                loaded_data = pd.read_csv(item[1][1])
                loaded_data.to_excel(writer, sheet_name=item[0])
                logging.info("Writing %s to ETRS" , item[0])

            elif not os.path.exists(item[1][1]):
                logging.critical(f"{item[0]} File not found")
                pass
        

        logging.info("Saving Master...")
        # DEBUGGING PATH
        book.save(fr"{BASE_PATH}\ETRS\\ETRS Master\\ETRS " + str(date.today()) + ".xlsx")
        # REAL PATH
        # book.save(fr"{BASE_PATH}\ETRS\\ETRS " + str(date.today()) + ".xlsx")
        logging.info(r"Master Saved!")

def refresh_excel_values(path):
    app = xl.App(visible=False)
    book = app.books.open(path)
    book.save()
    app.kill()


def tableify_etrs():
    """ Manipulates the ETRS in order to make it easier to report stats on"""

    today = date.today()
    bom_export_path = r"\BOM\BOM Exports\BOM Export "
    today_bom_format = str(today.strftime('%Y%m%d'))
    workflow_path = r"\BOM\Upchain Custom Reports\EBOM Reports\eBOM Workflow Report "
    todays_etrs = fr"{BASE_PATH}\ETRS\\ETRS Master\\ETRS " + str(date.today()) + ".xlsx"
    workflow_file = fr"{BASE_PATH}{workflow_path}{today_bom_format}.xlsx"
    today_bom_source = fr"{BASE_PATH}{bom_export_path}{today_bom_format}.xlsx"

    refresh_excel_values(todays_etrs) # refresh ETRS values
    


    etrs_df = pd.read_excel(r"S:\PDM Files\P1 - Mustang\ETRS\ETRS Master\ETRS 2022-06-29(1).xlsx", sheet_name="BTRS", skiprows = 4)
    workflow_df = pd.read_excel(r"S:\PDM Files\P1 - Mustang\BOM\Upchain Custom Reports\EBOM Reports\eBOM Workflow Report " + str(today.strftime('%Y%m%d') + ".xlsx"))
   # bom_df = pd.read_excel(today_bom_source, sheet_name = "Raw BOM")


    etrs_df.rename(columns={ 
        etrs_df.columns[22]: "Reqs Item Name",
        etrs_df.columns[23]: "Reqs Item Description",
        etrs_df.columns[24]: "Reqs Quantity",
        etrs_df.columns[25]: "Reqs Type",
        etrs_df.columns[26]: "Reqs Mass (grams)",
        etrs_df.columns[27]: "Reqs Revision Note",
        etrs_df.columns[28]: "Reqs Material",
        etrs_df.columns[29]: "Reqs Finish",
        etrs_df.columns[30]: "Reqs Safety Critical",
        etrs_df.columns[31]: "Reqs 3D Model",
        etrs_df.columns[32]: "Reqs Fixings and Torque Value Required",
        etrs_df.columns[33]: "Reqs Fixings and Torque Value Complete",
        etrs_df.columns[34]: "Reqs 2D Drawings Required",
        etrs_df.columns[35]: "Reqs 2d Drawings Complete",
        etrs_df.columns[40]: "Reqs Supplier Nomination Status",
        etrs_df.columns[41]: "Reqs Manufacturer",
        etrs_df.columns[42]: " - ",
        etrs_df.columns[43]: "Reqs Development Lead Time",
        etrs_df.columns[44]: "Reqs Tool Lead Time",
        etrs_df.columns[45]: "Reqs Production Lead Time",
        etrs_df.columns[47]: "Reqs PPAP Complete",
        etrs_df.columns[48]: "Reqs PPAP Timing",
        etrs_df.columns[49]: "Reqs SIP Timing",
        etrs_df.columns[51]: "Item Number"
    }, inplace=True)

    selected_columns_df = etrs_df.filter([
        "Item Number",
        "Treepath", 
        "Revision Note", 
        "Function Group", 
        "Status",
        "Reqs Item Name",
        "Reqs Item Description",
        "Reqs Quantity",
        "Reqs Type",
        "Reqs Mass (grams)",
        "Reqs Revision Note",
        "Reqs Material",
        "Reqs Finish",
        "Reqs Safety Critical",
        "Reqs 3D Model",
        "Reqs Fixings and Torque Value Required",
        "Reqs Fixings and Torque Value Complete",
        "Reqs 2D Drawings Required",
        "Reqs 2d Drawings Complete",
        "Reqs Supplier Nomination Status",
        "Reqs Manufacturer",
        "Reqs Development Lead Time",
        "Reqs Tool Lead Time",
        "Reqs Production Lead Time",
        "Reqs PPAP Complete",
        "Reqs PPAP Timing",
        "Reqs SIP Timing",
        "New Part from Yesterday",
        "New Part from Monday",
        ])

    merged_df = pd.merge(selected_columns_df, workflow_df, how="outer", on="Item Number")
       
    gateway_conditions = [
        (merged_df["Workflow"] == "PDM Release") & (merged_df["Revision Note_x"].str.contains('G2', na = False)),
        (merged_df["Workflow"].str.contains("G2", na = False)),
        (merged_df["Workflow"].str.contains("G3", na = False)),
        (merged_df["Workflow"].str.contains("G1", na = False)),
        (merged_df["Revision Note_x"].str.contains("G3", na = False)),
        (merged_df["Revision Note_x"].str.contains("G2", na = False)),
        (merged_df["Revision Note_x"].str.contains("G1", na = False)),
        (merged_df["Revision Note_y"].str.contains("G1", na = False)),
        (merged_df["Revision Note_y"].str.contains("Initial", na = False)),
    ]

    merged_df["Gateway"] = np.select(gateway_conditions, 
                                    ["G2 - Release", "G2 - Release", "G3 - Release", 
                                     "G1 - Release", "G3 - Release", "G2 - Release",
                                     "G1 - Release", "G1 - Release", "G1 - Release"],
                                     default = "Unreleased")
   

    merged_df["Length of Item Name"] = merged_df["Item Name"].astype(str).map(len)

    fixing_conditions = [
        (merged_df["Item Name"].str.contains("PHANTOM", na = False)),
        (merged_df["Length of Item Name"] > 10),
        # TODO Add item name ends in 8s
        ]


    merged_df["Fixing"] = np.select(fixing_conditions, ["No", "Yes"])

    merged_df.to_csv(fr"{BASE_PATH}\ETRS\ETRS Master\output.csv")

    prepping_flat_df = merged_df.filter(["Item Name", "Item Description", "Quantity", "Part Type", "Item Number"])
    prepping_flat_df.set_index("Item Number")

    print(prepping_flat_df.info())
    print(prepping_flat_df.head())
    test_df = prepping_flat_df.loc[(prepping_flat_df['Part Type'] == "Purchased Item") | (prepping_flat_df['Part Type'] == "Purchased ElectroMechanical Part") | (prepping_flat_df['Part Type'] == "Purchased Electrical Part") | (prepping_flat_df['Part Type'] == "Purchased Mechanical Part")]
    
    
    
    output = test_df.pivot_table(index = ["Item Name", "Item Description", "Part Type", "Item Number"], values="Quantity")
    
    print(output.info())
    print(output["Part Type"])
    # output.to_csv(fr"{BASE_PATH}\ETRS\ETRS Master\output2.csv")

def get_new_parts(today_bom, compared_bom):
    new_parts_df = pd.concat([today_bom, compared_bom])

    new_parts_df.drop_duplicates(subset= "Item Name", keep = False, inplace = True)
    new_parts_df["Estimated Release Week"] = np.nan
    new_parts_df["Estimated Actual Date"] = np.nan
    new_parts_df["Comments"] = np.nan
    new_parts_df["PPAP Planned Timing"] = np.nan
    new_parts_df["PPAP Actual Timing"] = np.nan
    new_parts_df["Logistics Estimated Timing"] = np.nan
    new_parts_df["Logistics Planned Timing"] = np.nan
    new_parts_df["Logistics Actual Timing"] = np.nan
    return new_parts_df

def save_new_parts(dataframe, filename):
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(dataframe, index=True, header=True):
        ws.append(r)

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.save(fr"S:\PDM Files\P1 - Mustang\ETRS\New Parts\{filename}\{filename} New Parts {today_bom_format}.xlsx")

def update_timing_sheet():
    wb = load_workbook(r"S:\PDM Files\P1 - Mustang\ETRS\New Parts\Timing Sheet.xlsx")
    ws = wb.active
    df = get_new_parts(today_bom_df, monday_bom_df)
    new_parts_list = df.values.tolist()
    for data in new_parts_list:
        ws.append(data)
    wb.save(filename = r"S:\PDM Files\P1 - Mustang\ETRS\New Parts\Timing Sheet.xlsx")

def main():
    """ Wrapper function for running the major elements of the script in order
    """

    logging.info("\n\n")
    logging.info("Updating ETRS...")
    
    # write_to_finance_update_csv(
    #     "1OZemQa88tV9a4_-21oaAQnt5mbAo1Y7WLTXCDM7jIoE",
    #     fr"{BASE_PATH}\ETRS\DataFiles\\Finance " + str(date.today()) + ".csv"
    #     )
    # DEBUG Cancel archiving when debugging
    # excel_archiver()
    write_to_etrs()
    save_new_parts(get_new_parts(today_bom_df, yesterday_bom_df), "Daily")
    save_new_parts(get_new_parts(today_bom_df, monday_bom_df), "Weekly")
    update_timing_sheet()
    # tableify_etrs()
 
    logging.info("Update Successful")

if __name__ == "__main__":
    main()